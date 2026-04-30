"""
File    : backend/services/report_service.py
Author  : 김민정
WBS     : REP-01, REP-02
Create  : 2026-04-04
Description :

Modification History :
    - 2026-04-06 (김민정) : AI 검토 데이터를 바탕으로 PDF 감리 보고서 및 통계형 Excel 파일 생성(더미데이터로 생성됨)
"""

import os
import cv2
import shutil
import numpy as np
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, Image, KeepTogether
from reportlab.lib import colors
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.pagesizes import A4
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.ttfonts import TTFont
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment

# 한글 폰트 설정
def register_korean_font():
    font_paths = [
        "NanumGothic.ttf",                     # 1. 앱 내 폰트
        "C:/Windows/Fonts/malgun.ttf",         # 2. Windows 맑은 고딕
        "C:/Windows/Fonts/batang.ttc",         # 3. Windows 바탕 (ttc 지원 여부에 따라 다를 수 있음)
        "/usr/share/fonts/truetype/nanum/NanumGothic.ttf", # 4. Linux (Ubuntu 등)
    ]
    
    for path in font_paths:
        try:
            if os.path.exists(path) or path == "NanumGothic.ttf":
                pdfmetrics.registerFont(TTFont('KoreanFont', path))
                return 'KoreanFont'
        except:
            continue
    return 'Helvetica'

FONT_NAME = register_korean_font()

class FormalReportGenerator:
    def __init__(self, output_path="formal_audit_report.pdf"):
        self.output_path = output_path
        self.styles = getSampleStyleSheet()
        self._set_custom_styles()
        self.temp_image_dir = "temp_captures"
        os.makedirs(self.temp_image_dir, exist_ok=True)

    def _set_custom_styles(self):
        """리포트 스타일 정의 (colors 속성 활용)"""
        self.styles.add(ParagraphStyle(name='KoNormal', fontName=FONT_NAME, fontSize=10, leading=14))
        self.styles.add(ParagraphStyle(name='KoTitle', parent=self.styles['Title'], fontName=FONT_NAME, fontSize=24, spaceAfter=40))
        self.styles.add(ParagraphStyle(name='KoSubtitle', fontName=FONT_NAME, fontSize=13, leading=14))
        self.styles.add(ParagraphStyle(name='KoHeading2', parent=self.styles['Heading2'], fontName=FONT_NAME, fontSize=14, spaceBefore=25, spaceAfter=15, textColor=colors.navy))
        self.styles.add(ParagraphStyle(name='TableHeaderText', fontName=FONT_NAME, fontSize=11, alignment=1, textColor=colors.white))
        self.styles.add(ParagraphStyle(name='TableText', fontName=FONT_NAME, fontSize=9, alignment=1))
        self.styles.add(ParagraphStyle(name='CenterText', fontName=FONT_NAME, fontSize=9, alignment=1))

    def _get_severity_html(self, severity):
        """심각도 아이콘 태그 (PDF 호환 유색 원)"""
        if severity == "Critical":
            return '<font color="red">●</font>'
        elif severity == "Major":
            return '<font color="orange">●</font>'
        else:
            return '<font color="gold">●</font>'

    def _capture_violation_zone(self, drawing_path, violation_id, coords, padding=150):
        """도면 위반 구역 캡처 및 강조"""
        img = cv2.imread(drawing_path)
        if img is None:
            placeholder_w, placeholder_h = 400, 300
            img_placeholder = np.full((placeholder_h, placeholder_w, 3), (200, 200, 200), dtype=np.uint8)
            text = "Image Not Found"
            font = cv2.FONT_HERSHEY_SIMPLEX
            (text_w, text_h), _ = cv2.getTextSize(text, font, 0.6, 2)
            cv2.putText(img_placeholder, text, ((placeholder_w - text_w) // 2, (placeholder_h + text_h) // 2), font, 0.6, (50, 50, 50), 2)
            save_path = os.path.join(self.temp_image_dir, f"{violation_id}_error.png")
            cv2.imwrite(save_path, img_placeholder)
            return save_path

        h_img, w_img, _ = img.shape
        x, y, w, h = coords['x'], coords['y'], coords['w'], coords['h']
        (text_w, text_h), _ = cv2.getTextSize(violation_id, cv2.FONT_HERSHEY_SIMPLEX, 0.7, 2)

        cv2.rectangle(img, (x, y), (x + w, y + h), (0, 0, 255), 3)
        cv2.rectangle(img, (x, y - 40), (x + text_w + 20, y), (0, 0, 255), -1)
        cv2.putText(img, violation_id, (x + 10, y - 10), cv2.FONT_HERSHEY_SIMPLEX, 0.7, (255, 255, 255), 2)
        
        x1, y1 = max(0, x - padding), max(0, y - padding)
        x2, y2 = min(w_img, x + w + padding), min(h_img, y + h + padding)
        crop_img = img[y1:y2, x1:x2]
        
        save_path = os.path.join(self.temp_image_dir, f"{violation_id}_crop.png")
        cv2.imwrite(save_path, crop_img)
        return save_path

    def create_report(self, data):
            """최종 PDF 감리 리포트 생성 (요약 바로 뒤에 상세 내역 배치)"""
            doc = SimpleDocTemplate(self.output_path, pagesize=A4, 
                                    topMargin=40, bottomMargin=40, leftMargin=40, rightMargin=40)
            elements = []

            # [1] 문서 제목
            elements.append(Paragraph(data.get('report_title', 'AI 기반 도면 검토 감리 리포트'), self.styles['KoTitle']))

            # [2] 기본 정보 테이블 (상단 헤더)
            sd = data['summary']
            agents_display = ", ".join(sd.get('agents', [])) if isinstance(sd.get('agents'), list) else str(sd.get('agents', '-'))
            header_data = [
                [Paragraph("<b>프로젝트명</b>", self.styles['TableText']), Paragraph(sd['project_name'], self.styles['KoNormal']),
                Paragraph("<b>검토일시</b>", self.styles['TableText']), Paragraph(sd['check_date'], self.styles['TableText'])],
                [Paragraph("<b>도면번호</b>", self.styles['TableText']), Paragraph(sd['drawing_no'], self.styles['KoNormal']),
                Paragraph("<b>사용 에이전트</b>", self.styles['TableText']), Paragraph(agents_display, self.styles['TableText'])]
            ]
            h_table = Table(header_data, colWidths=[70, 185, 70, 185])
            h_table.setStyle([
                ('GRID', (0,0), (-1,-1), 0.5, colors.black), 
                ('BACKGROUND', (0,0), (0,-1), colors.whitesmoke), 
                ('VALIGN', (0,0), (-1,-1), 'MIDDLE')
            ])
            elements.append(h_table)
            elements.append(Spacer(1, 20)) # 간격을 약간 줄여 공간 확보

            # [3] 섹션 1: 도면 검토 결과 요약
            elements.append(Paragraph("■ 1. 도면 검토 결과 요약", self.styles['KoHeading2']))
            list_data = [[Paragraph(f"<b>{h}</b>", self.styles['TableHeaderText']) for h in ["ID", "심각도", "규격 조문", "좌표", "설명 및 권장 조치"]]]
            for v in data['violations']:
                icon = self._get_severity_html(v['severity'])
                list_data.append([
                    Paragraph(v['id'], self.styles['TableText']),
                    Paragraph(f"{icon} {v['severity']}", self.styles['TableText']),
                    Paragraph(v['rule'], self.styles['TableText']),
                    Paragraph(f"({v['coords']['x']}, {v['coords']['y']})", self.styles['TableText']),
                    Paragraph(f"{v['desc']}<br/><b>[조치]</b> {v.get('recommendation', '-')}", self.styles['KoNormal'])
                ])
            l_table = Table(list_data, colWidths=[60, 55, 90, 65, 225])
            l_table.setStyle([
                ('GRID', (0,0), (-1,-1), 0.5, colors.black), 
                ('BACKGROUND', (0,0), (-1,0), colors.navy), 
                ('VALIGN', (0,0), (-1,-1), 'MIDDLE')
            ])
            elements.append(l_table)
            
            elements.append(Spacer(1, 30)) # 섹션 간 간격

            # [4] 섹션 2: 위반 위치 상세 첨부 (요약 테이블 바로 밑에서 시작)
            elements.append(Paragraph("■ 2. 위반 위치 상세 첨부", self.styles['KoHeading2']))
            TOTAL_WIDTH = 515

            for i, v in enumerate(data['violations']):
                v_el = []
                capture_path = self._capture_violation_zone(data['original_drawing'], v['id'], v['coords'])
                icon = self._get_severity_html(v['severity'])
                
                # (1) 소제목 및 심각도
                v_el.append(Paragraph(f"<b>[{v['id']}] {v['rule']} 위반</b>", self.styles['KoSubtitle']))
                sev_box = Table([[Paragraph(f"심각도: {icon} {v['severity']}", self.styles['CenterText'])]], colWidths=[90], hAlign='RIGHT')
                sev_box.setStyle([('BOX', (0,0), (-1,-1), 0.5, colors.grey), ('VALIGN', (0,0), (-1,-1), 'MIDDLE'), ('PADDING', (0,0), (-1,-1), 5)])
                v_el.append(sev_box)
                v_el.append(Spacer(1, 10))

                # (2) 상세 내역 텍스트 표 (이미지 위로 올림)
                detail_data = [
                    [Paragraph("<b>위반내용</b>", self.styles['TableText']), Paragraph(v['desc'], self.styles['KoNormal'])],
                    [Paragraph("<b>권장조치</b>", self.styles['TableText']), Paragraph(v.get('recommendation', '-'), self.styles['KoNormal'])]
                ]
                detail_table = Table(detail_data, colWidths=[80, 425])
                detail_table.setStyle([
                    ('GRID', (0,0), (-1,-1), 0.5, colors.lightgrey),
                    ('BACKGROUND', (0,0), (0,-1), colors.whitesmoke),
                    ('VALIGN', (0,0), (-1,-1), 'TOP'),
                    ('PADDING', (0,0), (-1,-1), 6)
                ])
                v_el.append(detail_table)
                v_el.append(Spacer(1, 10))

                # (3) 캡처 이미지 (텍스트 표 밑으로 배치)
                if capture_path and os.path.exists(capture_path):
                    # 높이를 180~200정도로 조절하면 첫 페이지에 요약+상세1건이 들어갈 확률이 높습니다.
                    img_obj = Image(capture_path, width=TOTAL_WIDTH, height=180, kind='proportional')
                    img_box = Table([[img_obj], [Paragraph(f"CAD Position: {v['coords']['x']}, {v['coords']['y']}", self.styles['CenterText'])]], colWidths=[TOTAL_WIDTH])
                    img_box.setStyle([('BOX', (0,0), (-1,-1), 0.5, colors.grey), ('ALIGN', (0,0), (-1,-1), 'CENTER'), ('BOTTOMPADDING', (0,0), (-1,-1), 8)])
                    v_el.append(img_box)

                v_el.append(Spacer(1, 25)) 

                # KeepTogether를 사용하여 상세 내역 한 세트가 페이지 중간에 잘리지 않게 함
                elements.append(KeepTogether(v_el))

            # [5] 리포트 빌드
            doc.build(elements)
            print(f"리포트 통합 생성 완료: {self.output_path}")
            self._cleanup_temp_images()

    def _cleanup_temp_images(self):
        try:
            if os.path.exists(self.temp_image_dir):
                shutil.rmtree(self.temp_image_dir)
                os.makedirs(self.temp_image_dir, exist_ok=True)
                print(f"임시 이미지 정리 완료")
        except Exception as e:
            print(f"오류 발생: {e}")

    def create_excel_report(self, data, excel_path):
            """요구사항 REP-02 대응: Excel 리포트 고도화"""
            wb = Workbook()

            # --- [1] 위반 항목 시트 (상세 내역) ---
            ws_detail = wb.active
            ws_detail.title = "위반 항목 목록"

            headers = ["ID", "심각도", "규격 조문", "위치", "설명", "권장 조치"]
            ws_detail.append(headers)

            # 심각도별 색상 정의 (PatternFill)
            fills = {
                "Critical": PatternFill(start_color="FFCCCC", end_color="FFCCCC", fill_type="solid"), # 연한 빨강
                "Major": PatternFill(start_color="FFE5CC", end_color="FFE5CC", fill_type="solid"),    # 연한 주황
                "Minor": PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")     # 연한 노랑
            }

            for v in data['violations']:
                row_data = [
                    v['id'], v['severity'], v['rule'],
                    f"({v['coords']['x']}, {v['coords']['y']})",
                    v['desc'], v.get('recommendation', '')
                ]
                ws_detail.append(row_data)
                
                # [색상 적용] 현재 행의 '심각도' 셀에 색상 입히기
                current_row = ws_detail.max_row
                severity = v['severity']

                if severity in fills:
                    ws_detail.cell(row=current_row, column=2).fill = fills[severity]

                for col_num in [5, 6]:
                    ws_detail.cell(row=current_row, column=col_num).alignment = Alignment(wrap_text=True, vertical='top')

            # [필터 적용] 모든 데이터에 필터 추가
            ws_detail.auto_filter.ref = f"A1:F{ws_detail.max_row}"

            # --- [2] 요약 시트 생성 (REP-02: 심각도/에이전트/구역별 건수) ---
            ws_summary = wb.create_sheet("요약 정보")
            sd = data['summary']

            # 제목 스타일 설정
            title_font = Font(bold=True, size=12)
            ws_summary.append(["■ 통계 요약"])
            ws_summary["A1"].font = title_font
            ws_summary.append(["항목", "내용"])
            ws_summary.append(["프로젝트명", sd['project_name']])
            ws_summary.append(["검토 도면", sd['drawing_no']])
            ws_summary.append([]) # 빈 줄

            # 1. 심각도별 위반 건수
            ws_summary.append(["[심각도별 위반 건수]"])
            ws_summary[f"A{ws_summary.max_row}"].font = Font(bold=True)

            # 각 등급별 데이터 행 추가 및 색상 적용
            severity_rows = [
                ("Critical", sd.get('sev_critical', 0), "FFCCCC"), # 연한 빨강
                ("Major", sd.get('sev_major', 0), "FFE5CC"),       # 연한 주황
                ("Minor", sd.get('sev_minor', 0), "FFFFCC")        # 연한 노랑
            ]

            for label, count, color in severity_rows:
                ws_summary.append([label, count])
                curr_row = ws_summary.max_row
                # A열(항목명)에 색상 배경과 볼드체 적용
                ws_summary.cell(row=curr_row, column=1).fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
                ws_summary.cell(row=curr_row, column=1).font = Font(bold=True)
                # B열(숫자)은 우측 정렬
                ws_summary.cell(row=curr_row, column=2).alignment = Alignment(horizontal='right')

            # 합계행 강조
            ws_summary.append(["합계", sd['total_violations']])
            sum_row = ws_summary.max_row
            ws_summary.cell(row=sum_row, column=1).font = Font(bold=True)
            ws_summary.cell(row=sum_row, column=2).font = Font(bold=True)
            ws_summary.append([]) # 빈 줄

            # 2. 에이전트별 위반 건수
            ws_summary.append(["[에이전트별 위반 현황]"])
            ws_summary.cell(row=ws_summary.max_row, column=1).font = Font(bold=True)
            ws_summary[f"A{ws_summary.max_row}"].font = Font(bold=True)
            
            # 에이전트별 카운트 로직
            agent_stats = {}
            for v in data['violations']:
                a_name = v.get('agent', '미지정')
                agent_stats[a_name] = agent_stats.get(a_name, 0) + 1
                
            for a_name, count in agent_stats.items():
                ws_summary.append([a_name, count])
                ws_summary.cell(row=ws_summary.max_row, column=2).alignment = Alignment(horizontal='right')
            ws_summary.append([])

            # 3. 구역별 위반 건수
            ws_summary.append(["[구역별 위반 현황]"])
            ws_summary[f"A{ws_summary.max_row}"].font = Font(bold=True)
            
            # 구역별 카운트 로직
            zone_stats = {}
            for v in data['violations']:
                z_name = v.get('zone', '전체 구역')
                zone_stats[z_name] = zone_stats.get(z_name, 0) + 1
                
            for z_name, count in zone_stats.items():
                ws_summary.append([z_name, count])
                ws_summary.cell(row=ws_summary.max_row, column=2).alignment = Alignment(horizontal='right')

            # --- [3] 표 디자인 (셀 너비 등) ---
            for ws in [ws_detail, ws_summary]:
                for col in ws.columns:
                    max_length = 0
                    column = col[0].column_letter
                    for cell in col:
                        try:
                            if len(str(cell.value)) > max_length:
                                max_length = len(str(cell.value))
                        except: pass
                    adjusted_width = (max_length + 2) * 1.2
                    ws.column_dimensions[column].width = min(adjusted_width, 50) # 최대 50자 제한

            wb.save(excel_path)
            print(f"REP-02: Excel 생성 완료 ({excel_path})")